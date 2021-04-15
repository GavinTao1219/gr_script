from collections import defaultdict
from pprint import pprint

import openpyxl
import os
import const

# {name:{date:{project: hour}}}
mingxi = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

# {name: {project: hour}}
total = defaultdict(lambda: defaultdict(float))


def parse_excel(excel_path):
    sheet = openpyxl.load_workbook(excel_path).active
    name = sheet.cell(const.NAME_CELL_ROW, const.NAME_CELL_COL).value
    for big_line_index in range(3):
        for date_index_in_big_line in range(const.DAYS_PER_BIG_LINE):
            date_row = const.DATE_START_CELL_ROW + (const.DATE_LINE_OCCUPIED + const.PROJECT_COUNT) * big_line_index
            date_col = const.DATE_START_CELL_COL + 3 * date_index_in_big_line
            date = sheet.cell(date_row, date_col).value
            for project_index in range(const.PROJECT_COUNT):
                project_name = sheet.cell(date_row + const.DATE_LINE_OCCUPIED + project_index, date_col).value
                project_hour = sheet.cell(date_row + const.DATE_LINE_OCCUPIED + project_index, date_col + 2).value
                if name is not None and date is not None and project_name is not None and project_hour is not None:
                    mingxi[name][date][project_name.upper()] += project_hour
                    total[name][project_name.upper()] += project_hour


def write_result_mingxi(output_filename):
    book = openpyxl.Workbook()
    name_index = 0
    for name in mingxi:
        sheet = book.worksheets[name_index]
        sheet.title = name
        for date_index, date in enumerate(mingxi[name]):
            sheet.cell(date_index + 2, 1).value = date
            for project_name in const.PROJECT_LOCATION:
                sheet.cell(1, const.PROJECT_LOCATION[project_name]).value = project_name
            for project_name in mingxi[name][date]:
                sheet.cell(date_index + 2, const.PROJECT_LOCATION[project_name]).value = mingxi[name][date][project_name]
        book.create_sheet()
        name_index += 1
    book.save(output_filename)


def write_result_total(output_filename):
    book = openpyxl.Workbook()
    for name_index, name in enumerate(total):
        sheet = book.active
        sheet.cell(name_index + 2, 1).value = name
        for project_name in const.PROJECT_LOCATION:
            sheet.cell(1, const.PROJECT_LOCATION[project_name]).value = project_name
        for project_name in total[name]:
            sheet.cell(name_index + 2, const.PROJECT_LOCATION[project_name]).value = total[name][project_name]
    book.save(output_filename)


def main():
    root_dir = const.ROOT_DIR
    for file in os.listdir(root_dir):
        d = os.path.join(root_dir, file)
        if os.path.isdir(d):
            for excel in os.listdir(d):
                parse_excel(os.path.join(d, excel))
    pprint(dict(mingxi))
    pprint(dict(total))
    write_result_mingxi(const.MINGXI_FILENAME)
    write_result_total(const.TOTAL_FILENAME)


if __name__ == "__main__":
    main()
